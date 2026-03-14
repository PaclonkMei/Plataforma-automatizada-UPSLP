from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import transaction
from django.utils.timezone import now
from django.core.paginator import Paginator
from django.db.models import Avg, Max, Min, Count, Q
from django.contrib import messages

from .models import Calificacion
from usuarios.models import Usuario
from .forms import CalificacionForm
from .utils import enviar_notificacion_alumno, enviar_resumen_profesor  # ✅ NUEVO IMPORT
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO


@login_required
def registrar_calificacion(request):
    if request.user.rol != 'profesor':
        return redirect('no_autorizado')

    mensaje_exito = None
    
    if request.method == 'POST':
        form = CalificacionForm(request.POST)
        if form.is_valid():
            calificacion = form.save(commit=False)
            calificacion.profesor = request.user
            calificacion.save()
            
            # ✅ ENVIAR CORREO AL ALUMNO
            if enviar_notificacion_alumno(calificacion):
                mensaje_exito = f'✓ Calificación registrada y notificación enviada a {calificacion.alumno.get_full_name() or calificacion.alumno.username} ({calificacion.materia}: {calificacion.nota})'
            else:
                mensaje_exito = f'✓ Calificación registrada para {calificacion.alumno.get_full_name() or calificacion.alumno.username} ({calificacion.materia}: {calificacion.nota}). ⚠️ No se pudo enviar el correo.'
            
            form = CalificacionForm()  # Limpiar el formulario
    else:
        form = CalificacionForm()

    return render(request, 'calificaciones/registrar.html', {
        'form': form,
        'mensaje_exito': mensaje_exito
    })


def calificacion_exitosa(request):
    return HttpResponse("<h2>Calificación registrada exitosamente.</h2> <a href='/calificaciones/registrar/'>Registrar otra</a>")


def no_autorizado(request):
    return HttpResponse("<h2>No estás autorizado para acceder a esta página.</h2>")


@login_required
def ver_calificaciones(request):
    if request.user.rol != 'alumno':
        return redirect('no_autorizado')

    calificaciones = Calificacion.objects.filter(alumno=request.user).order_by('-fecha')
    return render(request, 'calificaciones/ver.html', {'calificaciones': calificaciones})


@login_required
def subir_calificaciones(request):
    """
    Vista para subir calificaciones desde CSV o Excel (.xlsx).
    Espera columnas: matricula, materia, calificacion
    """
    if request.user.rol != 'profesor':
        return redirect('no_autorizado')

    mensaje = None
    error = None

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if not archivo:
            return render(request, 'calificaciones/subir_calificaciones.html', {
                'mensaje': None,
                'error': 'No se seleccionó ningún archivo.'
            })

        nombre = archivo.name.lower()
        try:
            # Leer archivo con pandas
            if nombre.endswith('.csv'):
                df = pd.read_csv(archivo)
            elif nombre.endswith('.xlsx'):
                df = pd.read_excel(archivo, engine='openpyxl')
            else:
                return render(request, 'calificaciones/subir_calificaciones.html', {
                    'mensaje': None,
                    'error': 'Formato no soportado. Sube un archivo .csv o .xlsx'
                })

            # Normalizar encabezados
            df.columns = [str(c).strip().lower() for c in df.columns]

            # Validar columnas requeridas
            columnas_requeridas = ['matricula', 'materia', 'calificacion']
            faltantes = [c for c in columnas_requeridas if c not in df.columns]
            if faltantes:
                return render(request, 'calificaciones/subir_calificaciones.html', {
                    'mensaje': None,
                    'error': f"El archivo debe contener las columnas: {', '.join(columnas_requeridas)}."
                })

            # Procesar filas
            creados = actualizados = errores = 0
            errores_detalle = []
            calificaciones_guardadas = []  # ✅ NUEVO: Lista para enviar correos

            with transaction.atomic():
                for idx, fila in df.iterrows():
                    fila_num = idx + 2
                    try:
                        matricula_val = str(fila.get('matricula', '')).strip()
                        materia_val = str(fila.get('materia', '')).strip()
                        calif_raw = fila.get('calificacion', '')

                        if matricula_val == '' or materia_val == '' or calif_raw == '':
                            raise ValueError("Campos obligatorios vacíos (matricula/materia/calificacion).")

                        # Convertir calificación
                        try:
                            calif = float(str(calif_raw).replace(',', '.'))
                        except Exception:
                            raise ValueError(f"Calificación no numérica: '{calif_raw}'.")

                        if not (0.0 <= calif <= 100.0):
                            raise ValueError("Calificación fuera de rango (0–100).")

                        # Buscar alumno
                        try:
                            alumno = Usuario.objects.get(username=matricula_val, rol='alumno')
                        except Usuario.DoesNotExist:
                            raise ValueError(f"Alumno no encontrado: '{matricula_val}'.")

                        # Crear/actualizar calificación
                        objeto, created = Calificacion.objects.update_or_create(
                            alumno=alumno,
                            materia=materia_val,
                            defaults={'nota': calif, 'profesor': request.user}
                        )
                        
                        calificaciones_guardadas.append(objeto)  # ✅ NUEVO: Guardar para enviar correos
                        
                        if created:
                            creados += 1
                        else:
                            actualizados += 1

                    except Exception as e:
                        errores += 1
                        if len(errores_detalle) < 10:
                            errores_detalle.append(f"Fila {fila_num}: {e}")

            # ✅ ENVIAR CORREOS A LOS ALUMNOS
            correos_enviados = 0
            for calificacion in calificaciones_guardadas:
                if enviar_notificacion_alumno(calificacion):
                    correos_enviados += 1

            # ✅ ENVIAR RESUMEN AL PROFESOR
            enviar_resumen_profesor(request.user, creados, actualizados, errores)

            # Mensaje final
            if errores == 0:
                mensaje = f"✓ Importación exitosa. Creados: {creados}, Actualizados: {actualizados}. 📧 Correos enviados: {correos_enviados}. Se ha enviado un resumen a tu correo."
            else:
                detalle = " | ".join(errores_detalle)
                sufijo = "" if len(errores_detalle) == errores else f" (+{errores - len(errores_detalle)} más)"
                error = (f"Importación parcial. Creados: {creados}, Actualizados: {actualizados}, "
                         f"Errores: {errores}. 📧 Correos enviados: {correos_enviados}. Detalles: {detalle}{sufijo}")

        except Exception as e:
            error = f"Error al procesar el archivo: {str(e)}"

    return render(request, 'calificaciones/subir_calificaciones.html', {'mensaje': mensaje, 'error': error})


@login_required
def historial_profesor(request):
    """
    Vista de historial y estadísticas para profesores.
    Muestra todas las calificaciones con filtros y estadísticas.
    """
    if request.user.rol != 'profesor':
        return redirect('no_autorizado')

    # Obtener todas las calificaciones del profesor
    calificaciones = Calificacion.objects.filter(profesor=request.user).select_related('alumno').order_by('-fecha')

    # Filtros
    alumno_filtro = request.GET.get('alumno', '').strip()
    materia_filtro = request.GET.get('materia', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    # Aplicar filtros
    if alumno_filtro:
        calificaciones = calificaciones.filter(
            Q(alumno__username__icontains=alumno_filtro) |
            Q(alumno__first_name__icontains=alumno_filtro) |
            Q(alumno__last_name__icontains=alumno_filtro)
        )
    
    if materia_filtro:
        calificaciones = calificaciones.filter(materia__icontains=materia_filtro)
    
    if fecha_desde:
        calificaciones = calificaciones.filter(fecha__gte=fecha_desde)
    
    if fecha_hasta:
        calificaciones = calificaciones.filter(fecha__lte=fecha_hasta)

    # Calcular estadísticas
    stats = calificaciones.aggregate(
        promedio=Avg('nota'),
        maximo=Max('nota'),
        minimo=Min('nota'),
        total=Count('id')
    )
    
    # Calcular porcentaje de reprobación (nota < 60)
    reprobados = calificaciones.filter(nota__lt=60).count()
    total = stats['total'] or 1
    porcentaje_reprobacion = (reprobados / total) * 100 if total > 0 else 0

    # Paginación
    paginator = Paginator(calificaciones, 15)  # 15 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Obtener listas únicas para los filtros
    alumnos_unicos = Calificacion.objects.filter(profesor=request.user).values_list('alumno__username', flat=True).distinct()
    materias_unicas = Calificacion.objects.filter(profesor=request.user).values_list('materia', flat=True).distinct()

    context = {
        'page_obj': page_obj,
        'stats': stats,
        'porcentaje_reprobacion': round(porcentaje_reprobacion, 2),
        'reprobados': reprobados,
        'alumno_filtro': alumno_filtro,
        'materia_filtro': materia_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'alumnos_unicos': alumnos_unicos,
        'materias_unicas': materias_unicas,
    }

    return render(request, 'calificaciones/historial_profesor.html', context)


@login_required
def editar_calificacion(request, calificacion_id):
    """
    Permite al profesor editar una calificación que registró.
    """
    if request.user.rol != 'profesor':
        return redirect('no_autorizado')
    
    # Obtener la calificación y verificar que pertenece al profesor
    calificacion = get_object_or_404(Calificacion, id=calificacion_id, profesor=request.user)
    
    if request.method == 'POST':
        form = CalificacionForm(request.POST, instance=calificacion)
        if form.is_valid():
            calificacion_actualizada = form.save()
            
            # ✅ ENVIAR CORREO AL ALUMNO NOTIFICANDO LA ACTUALIZACIÓN
            if enviar_notificacion_alumno(calificacion_actualizada):
                messages.success(request, f'✓ Calificación actualizada y notificación enviada al alumno.')
            else:
                messages.success(request, f'✓ Calificación actualizada. ⚠️ No se pudo enviar el correo.')
            
            return redirect('calificaciones:historial_profesor')
    else:
        form = CalificacionForm(instance=calificacion)
    
    return render(request, 'calificaciones/editar_calificacion.html', {
        'form': form,
        'calificacion': calificacion
    })


@login_required
def eliminar_calificacion(request, calificacion_id):
    """
    Permite al profesor eliminar una calificación que registró.
    """
    if request.user.rol != 'profesor':
        return redirect('no_autorizado')
    
    # Obtener la calificación y verificar que pertenece al profesor
    calificacion = get_object_or_404(Calificacion, id=calificacion_id, profesor=request.user)
    
    if request.method == 'POST':
        alumno_nombre = calificacion.alumno.get_full_name() or calificacion.alumno.username
        materia = calificacion.materia
        calificacion.delete()
        messages.success(request, f'✓ Calificación de {alumno_nombre} en {materia} eliminada exitosamente.')
        return redirect('calificaciones:historial_profesor')
    
    return render(request, 'calificaciones/eliminar_calificacion_confirmacion.html', {
        'calificacion': calificacion
    })


@login_required
def exportar_historial_excel(request):
    """
    Exporta el historial de calificaciones a Excel con los filtros aplicados.
    """
    if request.user.rol != 'profesor':
        return redirect('no_autorizado')

    # Obtener calificaciones con los mismos filtros
    calificaciones = Calificacion.objects.filter(profesor=request.user).select_related('alumno').order_by('-fecha')

    # Aplicar filtros (mismo código que historial_profesor)
    alumno_filtro = request.GET.get('alumno', '').strip()
    materia_filtro = request.GET.get('materia', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    if alumno_filtro:
        calificaciones = calificaciones.filter(
            Q(alumno__username__icontains=alumno_filtro) |
            Q(alumno__first_name__icontains=alumno_filtro) |
            Q(alumno__last_name__icontains=alumno_filtro)
        )
    
    if materia_filtro:
        calificaciones = calificaciones.filter(materia__icontains=materia_filtro)
    
    if fecha_desde:
        calificaciones = calificaciones.filter(fecha__gte=fecha_desde)
    
    if fecha_hasta:
        calificaciones = calificaciones.filter(fecha__lte=fecha_hasta)

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial Calificaciones"

    # Estilos
    header_fill = PatternFill(start_color="0D099E", end_color="0D099E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = ['Matrícula', 'Alumno', 'Materia', 'Calificación', 'Fecha']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Datos
    for cal in calificaciones:
        ws.append([
            cal.alumno.username,
            cal.alumno.get_full_name() or cal.alumno.username,
            cal.materia,
            float(cal.nota),
            cal.fecha.strftime('%Y-%m-%d')
        ])

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    # Agregar estadísticas al final
    ws.append([])
    ws.append(['ESTADÍSTICAS'])
    
    stats = calificaciones.aggregate(
        promedio=Avg('nota'),
        maximo=Max('nota'),
        minimo=Min('nota'),
        total=Count('id')
    )
    
    reprobados = calificaciones.filter(nota__lt=60).count()
    total = stats['total'] or 1
    porcentaje_reprobacion = (reprobados / total) * 100 if total > 0 else 0

    ws.append(['Total de registros:', stats['total']])
    ws.append(['Promedio:', round(stats['promedio'], 2) if stats['promedio'] else 0])
    ws.append(['Calificación máxima:', float(stats['maximo']) if stats['maximo'] else 0])
    ws.append(['Calificación mínima:', float(stats['minimo']) if stats['minimo'] else 0])
    ws.append(['Reprobados:', reprobados])
    ws.append(['% Reprobación:', f"{round(porcentaje_reprobacion, 2)}%"])

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Crear respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="historial_calificaciones_{now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response